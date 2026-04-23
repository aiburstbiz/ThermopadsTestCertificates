import sys
import subprocess
import os
import time

# --- AUTO-REPAIR BLOCK ---
def install_missing():
    try:
        import xlrd
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd", "openpyxl"])

install_missing()

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import io
import datetime
import re

# --- 0. SESSION STATE ---
if 'generated' not in st.session_state:
    st.session_state.generated = False
if 'file_data' not in st.session_state:
    st.session_state.file_data = None
if 'file_name' not in st.session_state:
    st.session_state.file_name = ""

st.set_page_config(page_title="Thermopads Automator", layout="centered")

# --- 1. THE APPROVED THEME ---
st.markdown("""
    <style>
        .stApp { background-color: #f8f9fa; }
        h1, label, p, [data-testid="stWidgetLabel"] p { color: #C62828 !important; font-weight: bold !important; }
        div.stButton > button * { color: #000000 !important; font-weight: 900 !important; }
        div.stButton > button {
            background-color: #C62828 !important;
            border-radius: 5px; border: 2px solid #8E0000;
            width: 100%; height: 3.5em;
        }
    </style>
    """, unsafe_allow_html=True)

st.markdown("<h1 style='text-align: center;'>🛡️ Thermopads Test Certificate Automator</h1>", unsafe_allow_html=True)

# --- 2. UNIVERSAL SMART LOADER ---
def load_data_smart(file):
    try:
        raw_bytes = file.getvalue()
        fname = file.name.lower()
        if fname.endswith('.xlsx'):
            df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=str, engine='openpyxl')
        elif fname.endswith('.xls'):
            try:
                df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=str, engine='xlrd')
            except:
                text = raw_bytes.decode('utf-8', errors='ignore')
                df_raw = pd.read_csv(io.StringIO(text), header=None, sep=r'\t+', engine='python', dtype=str)
        else:
            text = raw_bytes.decode('utf-8', errors='ignore')
            sep = r'\t+' if '\t' in text else ','
            df_raw = pd.read_csv(io.StringIO(text), header=None, sep=sep, engine='python', dtype=str)
            
        header_row_index = 0
        keywords = ['orderno', 'primarysrno', 'channelid', 'heatingcable', 'sl.no', 'materialno']
        for i, row in df_raw.head(100).iterrows():
            row_vals = [str(v).strip().lower().replace(" ", "").replace("\t", "") for v in row.values if v is not None]
            if any(k in row_vals for k in keywords):
                header_row_index = i
                break
        
        df = df_raw.iloc[header_row_index:].copy()
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
        df.columns = [str(c).strip().replace('\t', '') for c in df.columns]
        df = df.loc[:, ~df.columns.str.contains('^Unnamed|^nan|^None|^$', na=False)]
        for col in df.columns:
            df[col] = df[col].astype(str).apply(lambda x: x.strip().replace('\t', '') if x != 'nan' else "")
        return df
    except Exception as e:
        return None

# --- 3. TEMPLATE INJECTION ENGINE ---
def generate_from_official_template(merged, log_placeholder, o_date, s_date, log_history):
    template_path = "schluter_template.xlsx"
    if not os.path.exists(template_path):
        st.error(f"CRITICAL: '{template_path}' missing!")
        return None, None

    wb = load_workbook(template_path)
    ws = wb.active 
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    side = Side(style='thin'); thin_b = Border(left=side, right=side, top=side, bottom=side)

    info = merged.iloc[0]
    ord_no = str(info.get('OrderNo', 'N/A'))
    ws['A8'] = f"Schluter's purchase order no. {ord_no}"
    ws['A9'] = f"                        Dt : {o_date}"
    ws['A10'] = f"Date of shipment : {s_date}"

    mod_col = 'CustomerCode' if 'CustomerCode' in merged.columns else 'MaterialNo'
    cable_col = [c for c in merged.columns if any(x in c.lower() for x in ['cable', 'channel', 'primary'])][0]

    def extract_last_3(s):
        if not s or str(s).lower() == 'nan': return 0
        nums = re.findall(r'\d+', str(s))
        if nums:
            combined = "".join(nums)
            return int(combined[-3:]) if len(combined) >= 3 else int(combined)
        return 0

    merged['sort_key'] = merged[mod_col].apply(extract_last_3)
    merged_sorted = merged.sort_values(by='sort_key', ascending=False).reset_index(drop=True)

    # --- UPDATED ROW SHIFTING LOGIC ---
    current_excel_row = 17
    previous_model = None

    for i, row in merged_sorted.iterrows():
        current_model = str(row[mod_col])
        
        # If the Model No changes, physically insert an empty row to push everything down
        if previous_model is not None and current_model != previous_model:
            ws.insert_rows(current_excel_row)
            # We clear borders on this new inserted row to ensure it's a clean gap
            for cell in ws[current_excel_row]:
                cell.border = Border()
            current_excel_row += 1

        ts = datetime.datetime.now().strftime('%H:%M:%S')
        log_history.append(f"[{ts}] MAPPING: {current_model} -> Row {current_excel_row}")
        log_placeholder.code("\n".join(log_history[-10:]))
        
        row_data = [row[mod_col], row[cable_col], row.get('Expectedminout',''), 
                    row.get('Expectedmaxout',''), row.get('Actualminout','')]
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_excel_row, column=c_idx, value=str(val)) 
            cell.border = thin_b
            cell.alignment = center
            
        previous_model = current_model
        current_excel_row += 1
            
    return wb, ord_no

# --- 4. MAIN INTERFACE ---
st.subheader("📅 Certificate Details")
col1, col2 = st.columns(2)
with col1: order_dt = st.text_input("Order Date", value=datetime.datetime.now().strftime("%d.%m.%Y"))
with col2: ship_dt = st.text_input("Shipment Date", value=datetime.datetime.now().strftime("%d.%m.%Y"))

st.write("---")
c1, c2 = st.columns(2)
with c1: qc_file = st.file_uploader("1. QC Report", type=["xlsx", "xls", "csv"])
with c2: pack_file = st.file_uploader("2. Packing Data", type=["xlsx", "xls", "csv"])

if qc_file and pack_file:
    if not st.session_state.generated:
        if st.button("🚀 GENERATE OFFICIAL CERTIFICATE"):
            with st.expander("📊 LIVE PROCESSING LOGS", expanded=True):
                log_placeholder = st.empty()
                log_history = [f"[{datetime.datetime.now().strftime('%H:%M:%S')}] STARTING ENGINE..."]
                log_placeholder.code("\n".join(log_history))
                
                df_qc = load_data_smart(qc_file)
                df_pack = load_data_smart(pack_file)

                if df_qc is not None and df_pack is not None:
                    qc_id_list = [c for c in df_qc.columns if any(x in c.lower() for x in ['channel', 'cable'])]
                    pk_id_list = [c for c in df_pack.columns if any(x in c.lower() for x in ['primary', 'cable'])]
                    
                    if qc_id_list and pk_id_list:
                        qc_id = qc_id_list[0]
                        pk_id = pk_id_list[0]
                        merged = df_qc.merge(df_pack, left_on=qc_id, right_on=pk_id, how='inner')

                        if not merged.empty:
                            final_wb, order_id = generate_from_official_template(merged, log_placeholder, order_dt, ship_dt, log_history)
                            if final_wb:
                                output = io.BytesIO(); final_wb.save(output)
                                st.session_state.file_data = output.getvalue()
                                st.session_state.file_name = f"COA_{order_id}.xlsx"
                                st.session_state.generated = True
                                st.rerun()
                        else: st.error("No serial numbers matched.")
                    else: st.error("Columns (Channel/Primary) not found.")
    else:
        st.success(f"✅ Success! Your certificate is ready.")
        st.download_button("📥 DOWNLOAD CERTIFICATE", st.session_state.file_data, st.session_state.file_name)
        if st.button("🔄 START NEW BATCH"): 
            st.session_state.generated = False
            st.rerun()